import openpyxl

def readExcel(filepath,label,value):
    workbook = openpyxl.load_workbook(filepath)
    print(workbook.sheetnames)
    sheet = workbook["Sheet1"]
    noOfRows = sheet.max_row
    noOfCol = sheet.max_column
    print(noOfRows)
    print(noOfCol)
    Label = []
    Header = []
    for i in range(1,noOfRows+1):
        Label.append(sheet.cell(i,1).value)
        #print(Label[i-1])
        if Label[i-1] == label:
            for j in range(1, noOfCol + 1):
                Header.append(sheet.cell(1, j).value)
                #print(Header[j - 1])
                if Header[j - 1] == value:
                    val = sheet.cell(i,j).value
                    break
            break
    return val

print(readExcel("C:\\Users\\Vignesh\\Cucumber_AutomationBytes\\Booking\\data.xlsx","Google","Url"))